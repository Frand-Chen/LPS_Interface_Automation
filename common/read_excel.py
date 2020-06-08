# -*- coding=utf-8 -*-

"""
该模块用来操作 excel 测试用例文件(.xlsx格式)
"""

import openpyxl
from openpyxl.styles import Font, colors


class TestData:
    """保存测试用例数据"""
    pass


class ReadExcel:

    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name

    def open(self):
        """打开 excel，选择表单"""
        self.work_book = openpyxl.load_workbook(self.file_name)
        self.sheet = self.work_book[self.sheet_name]

    def read_data_dict(self) -> list:
        """
        读取测试用例数据，保存在 Dict 中
        :return: 列表嵌套 Dict
        """
        # 先清除 result 列的数据
        self.delete_result_column()

        # 按行获取所有的单元格数据
        row_sheets = list(self.sheet.rows)
        # 获取表头行
        title = []
        for sheet in row_sheets[0]:
            title.append(sheet.value)
        # 遍历获取每条测试用例数据，并保存在 list 中
        cases = []
        for row in row_sheets[1:]:
            data = []
            for sheet in row:
                data.append(sheet.value)
            # 把每个 sheet 的数据以 Dict 保存
            data_dict = dict(zip(title, data))
            cases.append(data_dict)
        return cases

    def read_data_object(self) -> list:
        """
        读取用例数据，保存在对象中
        :return: 列表嵌套对象
        """
        # 先清除 result 列的数据
        self.delete_result_column()

        # 按行获取所有的单元格数据
        row_sheets = list(self.sheet.rows)
        # 获取表头行
        title = []
        for sheet in row_sheets[0]:
            title.append(sheet.value)
        # 遍历获取每条测试用例数据，并保存在 list 中
        cases = []
        for row in row_sheets[1:]:
            data = []
            for sheet in row:
                data.append(sheet.value)
            data_dict = dict(zip(title, data))
            case_obj = TestData()
            # 遍历字典中的数据，设置为对象的属性和属性值
            for key, value in data_dict.items():
                if key != None:
                    setattr(case_obj, key, value)
            # 添加对象到 cases 中
            cases.append(case_obj)
        return cases

    def save(self):
        """保存 excel 文件"""
        self.work_book.save(self.file_name)

    def close(self):
        """关闭 excel 文件"""
        self.work_book.close()

    def write_data(self, row, column, value, font_color=colors.BLACK, bold=False):
        """
        写入数据
        :param row: 写入行
        :param column: 写入列
        :param value: 写入数据
        :param font_color: 字体颜色, 默认 BLACK
        :param bold: True 字体加粗, False 字体不加粗
        """
        # 先清除 result 列的数据
        self.open()
        font = Font(color=font_color, bold=bold)
        # 写入数据
        self.sheet.cell(row=row, column=column).font = font
        self.sheet.cell(row=row, column=column, value=value)
        self.save()
        self.close()

    def delete_result_column(self):
        """清除 result 列的数据"""
        self.open()
        row_sheets = list(self.sheet.rows)
        result_column = 0
        for sheet in row_sheets[0]:
            result_column += 1
            if sheet.value == "result":
                break
        # 遍历所有测试用例，清除 result 列的数据
        case_row = 2
        for row in row_sheets[1:]:
            font = Font(color=colors.BLACK, bold=False)
            self.sheet.cell(row=case_row, column=result_column).font = font
            self.sheet.cell(row=case_row, column=result_column, value="")
            case_row += 1
        self.save()
        self.close()


if __name__ == '__main__':
    data = ReadExcel(r"E:\CodeLibrary\LPS_Interface_Automation\data\LPS_testcases.xlsx", "Session")
    # case = data.read_data_dict()
    case = data.read_data_object()
    for ele in case:
        print(ele.case_id, ele.title,ele.result)
    # data.write_data(row=5, column=10, value="test", font_color=colors.RED, bold=True)
    # data.delete_result_column()
